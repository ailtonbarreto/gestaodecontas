#----------------------------------------------------------------------------------------
#Adicionar Entrada

import streamlit as st
import pandas as pd
import openpyxl as xl


#----------------------------------------------------------------------------------------
#exibição de dados

st.set_page_config(layout="wide",page_title="Adicionar Entrada",initial_sidebar_state='collapsed',page_icon='✅')


tab1, tab2, tab3 = st.tabs(['Adicionar Entrada','Excluir Entrada','Editar uma Entrada'])

#----------------------------------------------------------------------------------------
#Dados

df = pd.read_excel("Gestão de contas.xlsx",sheet_name='A Receber')
pd.to_datetime(df["Data Emissão"])
df["Ano"] = df["Data Emissão"].dt.year
df["Mês"] = df["Data Emissão"].dt.month
df.sort_values("Data Emissão")
df["Ano"].astype(int)
df["Mês"].astype(int)
df['Valor'].astype(float)

dfexcluir = df
dfexcluir["Indice"] = df.index
dfselectbox = pd.read_excel("Gestão de contas.xlsx",sheet_name='Cadastro de Fornecedores')
#----------------------------------------------------------------------------------------

def determinar_mês(valor):
    meses = {
        1: "Jan",
        2: "Fev",
        3: "Mar",
        4: "Abr",
        5: "Mai",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Set",
        10: "Out",
        11: "Nov",
        12: "Dez"
    }
    return meses.get(valor)


df["Mês"] = df["Mês"].apply(determinar_mês)
df = df.drop(columns=["Data Emissão"])

#----------------------------------------------------------------------------------------
#Adicionar Entrada

with tab1:

    st.title("🟢 Adicionar Entrada",anchor=False)

    entrada_fornecedor = st.selectbox("Fornecedor",dfselectbox['Fornecedor'].unique())

    entrada_notafiscal = st.text_input("Nota Fiscal")

    entrada_dataemissao = st.date_input("Data Emissão","today",format= "DD/MM/YYYY")

    entrada_datavencimento = st.date_input("Data Vencimento","today",format= "DD/MM/YYYY")
   
    entrada_valor = st.number_input("Valor")

    entrada_status = st.selectbox("Status",["A RECEBER","RECEBIDO"])

    if st.button("ADICIONAR"):
    # Abrir o arquivo do Excel
        planilha = xl.load_workbook("Gestão de contas.xlsx")
        planilha = planilha["A Receber"]

    # Criar uma nova linha com os dados inseridos
        nova_linha = [entrada_fornecedor, entrada_notafiscal, entrada_dataemissao, entrada_datavencimento, entrada_valor, entrada_status]
    
    # Adicionar a nova linha à planilha
        planilha.append(nova_linha)

    # Salvar as alterações de volta no arquivo Excel
        planilha.parent.save("Gestão de contas.xlsx")
   
        st.success("Movimentação salva!")


#----------------------------------------------------------------------------------------
#Excluir Entrada
with tab2:
    st.title("🟢 Excluir Entrada",anchor=False)

    filtro_ano = st.selectbox("Ano",dfexcluir["Ano"].unique())
    filtro_mes = st.selectbox("Mês",dfexcluir["Mês"].unique())
    filtro_cliente = st.selectbox('Cliente',dfexcluir["Cliente"].unique())
    linha = st.number_input("Excluir linha",format="%.0f")

    dfexcluir = dfexcluir.query('Ano == @filtro_ano & Mês == @filtro_mes & Cliente == @filtro_cliente')

    if st.button("EXCLUIR ENTRADA"):
        planilha = xl.load_workbook("Gestão de contas.xlsx")
        sheet = planilha["A Receber"]
    
        sheet.delete_rows(int(linha) + 2)
    
        planilha.save("Gestão de contas.xlsx")
    
        st.success("Entrada Excluída Com Sucesso!")

    # dfexcluir = dfexcluir.drop(columns=['Indice'])
    st.table(dfexcluir)
#----------------------------------------------------------------------------------------
#Editar Entrada

with tab3:
    st.title("🟢 Editar Entrada",anchor=False)

    dfeditarentrada = df
    
    #Dados da linha editada
    filtro_y = st.selectbox('Ano da Movimentação',dfeditarentrada["Ano"].unique())
    filtro_m = st.selectbox('Mês da Movimentação',dfeditarentrada["Mês"].unique())
    filtro_c = st.selectbox('Buscar Cliente',dfeditarentrada["Cliente"].unique())
    filtro_index = st.number_input("Linha a Editar",format="%.0f")
    editar_status = st.selectbox('Novo Status',["A RECEBER","RECIBIDO"])
    
#----------------------------------------------------------------------------------------
#Dataframe filtrado

    dfeditarentrada = df.query('Ano == @filtro_y & Mês == @filtro_m & Cliente == @filtro_c')

    linhaeditada = filtro_index + 2

    coluna = 6

    dfeditar = dfeditarentrada.drop(columns="Indice")

    if st.button("SALVAR EDIÇÃO"):
        planilha = xl.load_workbook("Gestão de contas.xlsx")
        sheet = planilha["A Receber"]
    
        editar = sheet.cell(row=linhaeditada,column=coluna)
        editar.value = editar_status
        novo_valor = editar
    
        planilha.save("Gestão de contas.xlsx")
    
        st.success("Edição salva!")

    dfeditar["Valor"] = dfeditar["Valor"].apply(lambda x: f'R$ {x:.2f}')
    
    st.table(dfeditarentrada)


#------------------------------------------------------------------------------------------
#Esconder streamlit menus

hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

