#Adiciona Saída

import streamlit as st
import pandas as pd
import openpyxl as xl

#----------------------------------------------------------------------------------------
#exibição de dados

st.set_page_config(layout="wide",page_title="Adicionar Saída",initial_sidebar_state='collapsed',page_icon='✅')


tab1, tab2, tab3 = st.tabs(["Adicionar Saída","Excluir Saída","Editar Status de Saída"])


#----------------------------------------------------------------------------------------
#Dados

df = pd.read_excel("Gestão de contas.xlsx",sheet_name="A Pagar")
pd.to_datetime(df["Data Emissão"])
df["Ano"] = df["Data Emissão"].dt.year
df["Indice"] = df.index
df["Mês"] = df["Data Emissão"].dt.month
df.sort_values("Data Emissão")
df["Ano"].astype(int)
df["Mês"].astype(int)
df['Valor'].astype(float)

#----------------------------------------------------------------------------------------
#cadastro de fornecedores
dfcadastro = pd.read_excel("Gestão de contas.xlsx",sheet_name="Cadastro de Fornecedores")


#----------------------------------------------------------------------------------------



def determinar_mês(valor):
    meses = {
        1: "Jan",
        2: "Fev",
        3: "Mar",
        4: "Abr",
        5: "Mai",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Set",
        10: "Out",
        11: "Nov",
        12: "Dez"
    }
    return meses.get(valor)

#----------------------------------------------------------------------------------------
#dicionário classificar meses

classificar_meses ={
    'Jan': 1,
    'Fev': 2,
    'Mar': 3,
    'Abr': 4,
    'Mai': 5,
    'Jun': 6,
    'Jul': 7,
    'Ago': 8,
    'Set': 9,
    'Out': 10,
    'Nov': 11,
    'Dez': 12
        }


df["Mês"] = df["Mês"].apply(determinar_mês)
df = df.drop(columns=["Data Emissão"])

df['Ordem_Mês'] = df['Mês'].map(classificar_meses)
df = df.sort_values(by='Ordem_Mês',ascending = True).drop(columns=['Ordem_Mês'])

#----------------------------------------------------------------------------------------
#Adicionar Saída

with tab1:
    st.title("🔴 Adicionar Saída",anchor=False)

    entrada_data = st.date_input("Data","today",format= "DD/MM/YYYY")

    entrada_fornecedor = st.selectbox("Fornecedor",dfcadastro["Fornecedor"])
    
    entrada_categoria = dfcadastro.loc[dfcadastro['Fornecedor'] == entrada_fornecedor, 'Categoria'].iloc[0]
   
    entrada_valor = st.number_input("Valor")

    entrada_status = st.selectbox("Status",["PAGO","A PAGAR"])


    if st.button("ADICIONAR SAÍDA"):
    # Abrir o arquivo do Excel
        planilha = xl.load_workbook("Gestão de contas.xlsx")
        planilha = planilha["A Pagar"]

    # Criar uma nova linha com os dados inseridos
        nova_linha = [entrada_data, entrada_fornecedor, entrada_categoria, entrada_valor, entrada_status,"SAÍDA"]
    
    # Adicionar a nova linha à planilha
        planilha.append(nova_linha)

    # Salvar as alterações de volta no arquivo Excel
        planilha.parent.save("Gestão de contas.xlsx")
   
        st.success("Movimentação salva!")
    
#------------------------------------------------------------------------------------------
#Remover linha
    
with tab2:

    st.title("🔴 Excluir Saída",anchor=False)
#Indice da linha a ser removida
    dfdelete = df
    filtro_ano = st.selectbox("Ano",dfdelete["Ano"].unique())
    filtro_mes = st.selectbox("Mês",dfdelete["Mês"].unique())
    filtro_fornecedor = st.selectbox('Fornecedor',dfdelete["Fornecedor"].unique())
    linha = st.number_input("Excluir Linha",format="%.0f")

    dfdelete = dfdelete.query('Ano == @filtro_ano & Mês == @filtro_mes & Fornecedor == @filtro_fornecedor')

    if st.button("EXCLUIR SAÍDA"):
        planilha = xl.load_workbook("Gestão de contas.xlsx")
        sheet = planilha["A Pagar"]
    
        sheet.delete_rows(int(linha) + 2)
    
        planilha.save("Gestão de contas.xlsx")
    
        st.success("Saída Excluída Com Sucesso!")

    dfdelete = dfdelete.drop(columns=['Indice'])
    st.table(dfdelete)
    
#------------------------------------------------------------------------------------------ 
# Editar uma saída  

with tab3:
    st.title("🔴 Editar Status",anchor=False)

    dfeditar = df
    
    #Dados da linha editada
    filtro_y = st.selectbox('Ano da Movimentação',dfeditar["Ano"].unique())
    filtro_m = st.selectbox('Mês da Movimentação',dfeditar["Mês"].unique())
    filtro_f = st.selectbox('Buscar Fornecedor',dfeditar["Fornecedor"].unique())
    filtro_index = st.number_input("Linha a Editar",format="%.0f")
    editar_status = st.selectbox('Novo Status',["A PAGAR","PAGO"])
    
#----------------------------------------------------------------------------------------
#Dataframe filtrado

    dfeditar = df.query('Ano == @filtro_y & Mês == @filtro_m & Fornecedor == @filtro_f')

    linhaeditada = filtro_index + 2

    coluna = 5

    dfeditar = dfeditar.drop(columns="Indice")
#----------------------------------------------------------------------------------------
#Editar Saída

    if st.button("SALVAR EDIÇÃO"):
        planilha = xl.load_workbook("Gestão de contas.xlsx")
        sheet = planilha["A Pagar"]
    
        editar = sheet.cell(row=linhaeditada,column=coluna)
        editar.value = editar_status
        novo_valor = editar
    
        planilha.save("Gestão de contas.xlsx")
    
        st.success("Edição salva!")

    dfeditar["Valor"] = dfeditar["Valor"].apply(lambda x: f'R$ {x:.2f}')
    
    st.table(dfeditar)
    
#------------------------------------------------------------------------------------------
#Esconder streamlit menus

hide_st_style = """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
    """
st.markdown(hide_st_style, unsafe_allow_html=True)